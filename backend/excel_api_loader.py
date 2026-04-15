"""
Excel API Loader - Loads API configurations from Excel files
Supports batch API testing with data from spreadsheet
"""

import json
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional
from pydantic import BaseModel
import re

# ============================================================================
# MODELS
# ============================================================================

class ExcelAPIConfig(BaseModel):
    api_name: str
    method: str  # GET, POST, PUT, DELETE, PATCH
    endpoint: str
    description: str = ""
    headers: Dict[str, str] = {}
    params: Dict[str, Any] = {}  # Query parameters
    body: Dict[str, Any] = {}  # Request body (for POST, PUT, PATCH)
    expected_status: List[int] = [200]
    auth_type: str = "none"  # none, bearer, basic
    auth_token: str = ""

# ============================================================================
# EXCEL LOADER
# ============================================================================

class ExcelAPILoader:
    """
    Loads API test configurations from Excel files.
    
    Expected Excel columns:
    - API Name: Name of the API
    - Method: HTTP method (GET, POST, PUT, DELETE, PATCH)
    - Endpoint: API endpoint path
    - Description: Optional description
    - Headers: JSON string of headers (optional)
    - Params: JSON string of query parameters (optional)
    - Body: JSON string of request body (optional)
    - Expected Status: Comma-separated list of expected HTTP status codes
    - Auth Type: none, bearer, basic
    - Auth Token: Authentication token/credentials
    """
    
    REQUIRED_COLUMNS = {
        'api_name': 'API Name',
        'method': 'Method',
        'endpoint': 'Endpoint'
    }
    
    OPTIONAL_COLUMNS = {
        'description': 'Description',
        'headers': 'Headers',
        'params': 'Params',
        'body': 'Body',
        'expected_status': 'Expected Status',
        'auth_type': 'Auth Type',
        'auth_token': 'Auth Token'
    }
    
    @staticmethod
    def load_from_excel(file_path: str) -> List[ExcelAPIConfig]:
        """
        Load API configurations from Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of ExcelAPIConfig objects
        """
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            
            # Normalize column names
            df.columns = df.columns.str.strip().str.lower()
            
            apis = []
            for idx, row in df.iterrows():
                try:
                    api_config = ExcelAPILoader._parse_row(row)
                    apis.append(api_config)
                except Exception as e:
                    print(f"Error parsing row {idx + 2}: {e}")
                    continue
            
            return apis
        except Exception as e:
            raise Exception(f"Failed to load Excel file: {str(e)}")
    
    @staticmethod
    def _parse_row(row: pd.Series) -> ExcelAPIConfig:
        """Parse a single row from Excel into ExcelAPIConfig"""
        
        # Extract required fields
        api_name = str(row.get('api name', '')).strip()
        method = str(row.get('method', '')).strip().upper()
        endpoint = str(row.get('endpoint', '')).strip()
        
        if not api_name or not method or not endpoint:
            raise ValueError("Missing required fields: API Name, Method, or Endpoint")
        
        # Validate method
        valid_methods = ['GET', 'POST', 'PUT', 'DELETE', 'PATCH', 'HEAD', 'OPTIONS']
        if method not in valid_methods:
            raise ValueError(f"Invalid HTTP method: {method}")
        
        # Extract optional fields
        description = str(row.get('description', '')).strip()
        
        # Parse JSON fields
        headers = ExcelAPILoader._parse_json(row.get('headers', '{}'), 'Headers')
        params = ExcelAPILoader._parse_json(row.get('params', '{}'), 'Params')
        body = ExcelAPILoader._parse_json(row.get('body', '{}'), 'Body')
        
        # Parse expected status codes
        expected_status = ExcelAPILoader._parse_status_codes(
            row.get('expected status', '200')
        )
        
        # Auth fields
        auth_type = str(row.get('auth type', 'none')).strip().lower()
        auth_token = str(row.get('auth token', '')).strip()
        
        return ExcelAPIConfig(
            api_name=api_name,
            method=method,
            endpoint=endpoint,
            description=description,
            headers=headers,
            params=params,
            body=body,
            expected_status=expected_status,
            auth_type=auth_type,
            auth_token=auth_token
        )
    
    @staticmethod
    def _parse_json(value: Any, field_name: str = "Field") -> Dict[str, Any]:
        """Safely parse JSON string"""
        if pd.isna(value) or (isinstance(value, str) and value.strip() == ''):
            return {}
        
        try:
            if isinstance(value, dict):
                return value
            if isinstance(value, str):
                return json.loads(value)
            return {}
        except json.JSONDecodeError as e:
            raise ValueError(f"{field_name} is not valid JSON: {e}")
    
    @staticmethod
    def _parse_status_codes(value: Any) -> List[int]:
        """Parse status codes from comma-separated string"""
        if pd.isna(value):
            return [200]
        
        try:
            if isinstance(value, list):
                return [int(x) for x in value]
            if isinstance(value, str):
                return [int(x.strip()) for x in value.split(',') if x.strip().isdigit()]
            return [int(value)]
        except (ValueError, AttributeError):
            return [200]
    
    @staticmethod
    def create_sample_excel(output_path: str) -> None:
        """
        Create a sample Excel file with proper formatting using openpyxl.
        
        Args:
            output_path: Path where to save the sample Excel file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'APIs'
            
            # Headers
            headers = [
                'API Name', 'Method', 'Endpoint', 'Description', 
                'Headers', 'Params', 'Body', 'Expected Status', 'Auth Type', 'Auth Token'
            ]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Sample data
            sample_rows = [
                ['Get Users', 'GET', '/api/users', 'Fetch all users', 
                 '{"Content-Type": "application/json"}', '{}', '{}', '200', 'none', ''],
                ['Create User', 'POST', '/api/users', 'Create a new user',
                 '{"Content-Type": "application/json"}', '{}', '{"name": "John", "email": "john@example.com"}', '201', 'none', ''],
                ['Update User', 'PUT', '/api/users/123', 'Update user by ID',
                 '{"Content-Type": "application/json"}', '{}', '{"name": "John Updated"}', '200', 'bearer', 'your_bearer_token_here'],
                ['Delete User', 'DELETE', '/api/users/123', 'Delete user by ID',
                 '{"Content-Type": "application/json"}', '{}', '{}', '204', 'bearer', 'your_bearer_token_here'],
                ['Login User', 'POST', '/api/auth/login', 'User login endpoint',
                 '{"Content-Type": "application/json"}', '{}', '{"username": "user@test.com", "password": "pass123"}', '200', 'none', ''],
            ]
            
            for row in sample_rows:
                ws.append(row)
            
            # Auto-adjust column widths
            column_widths = [15, 10, 20, 20, 35, 15, 40, 18, 12, 25]
            for idx, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + idx)].width = width
            
            # Save file
            wb.save(output_path)
            print(f"Sample Excel file created at: {output_path}")
            
        except Exception as e:
            print(f"Error creating Excel file: {e}")
            import traceback
            traceback.print_exc()
            raise
